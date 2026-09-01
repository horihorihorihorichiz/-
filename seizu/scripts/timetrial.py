# -*- coding: utf-8 -*-
"""タイムトライアル記録表（xlsx）を作る。"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

ROWS = 200

MENU = [
    ('読み取り', 25, '問題文を読んで条件を書き出すまで'),
    ('エスキス', 60, '間取りを決めて下書きが固まるまで'),
    ('記述', 25, '計画の要点等をすべて書き終えるまで'),
    ('1階平面図兼配置図', 40, '1階の平面図と配置図'),
    ('2階平面図', 30, '2階の平面図'),
    ('3階平面図', 25, '3階の平面図'),
    ('床伏図兼小屋伏図', 35, '骨組みの図'),
    ('立面図', 20, '横から見た図'),
    ('部分詳細図1/20', 25, '外壁の断面'),
    ('面積表', 5, '各階と延べ面積'),
    ('通し（5時間）', 300, '本番と同じ流れを最初から最後まで'),
]

INK = '1F2937'
ACC = 'B03060'
HEAD = PatternFill('solid', fgColor='F3EFE7')
ACCF = PatternFill('solid', fgColor=ACC)
GREEN = PatternFill('solid', fgColor='D8F0DC')
RED = PatternFill('solid', fgColor='FBDDDD')
GREY = PatternFill('solid', fgColor='FAFAF8')
thin = Side(style='thin', color='D8D2C6')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def head(ws, row, labels, widths):
    for i, (lab, w) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = ACCF
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 24


def build():
    wb = Workbook()

    # ---------------- 目標タイム ----------------
    ws2 = wb.active
    ws2.title = '目標タイム'
    ws2['A1'] = '目標タイム表（この時間で描けるようになるのが目標）'
    ws2['A1'].font = Font(bold=True, size=14, color=INK)
    head(ws2, 3, ['種目', '目標(分)', '何をする時間か'], [22, 10, 44])
    for i, (nm, mn, ds) in enumerate(MENU):
        r = 4 + i
        ws2.cell(row=r, column=1, value=nm).border = BOX
        c = ws2.cell(row=r, column=2, value=mn)
        c.border = BOX
        c.alignment = Alignment(horizontal='center')
        ws2.cell(row=r, column=3, value=ds).border = BOX
    last = 3 + len(MENU)
    ws2.cell(row=last + 2, column=1,
             value='※ 図面だけの合計 ＝ 180分。読み取り・エスキス・記述を'
                   '入れて290分。5時間＝300分なので、余りは10分しかない。')
    ws2.cell(row=last + 2, column=1).font = Font(size=10, color='6B6459')

    # ---------------- 記録 ----------------
    ws = wb.create_sheet('記録')
    ws['A1'] = 'タイムトライアル記録'
    ws['A1'].font = Font(bold=True, size=15, color=INK)
    ws['A2'] = ('B列で種目を選ぶと目標分が自動で入ります。'
                'D列に実際にかかった分を入れてください。')
    ws['A2'].font = Font(size=10, color='6B6459')
    head(ws, 4, ['日付', '種目', '目標(分)', '実際(分)', '差(分)',
                 '何回目', 'できなかったこと・気づき'],
         [12, 22, 10, 10, 10, 9, 46])

    dv = DataValidation(type='list',
                        formula1='=目標タイム!$A$4:$A$%d' % last,
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)

    for r in range(5, 5 + ROWS):
        ws.cell(row=r, column=1).number_format = 'm"/"d'
        ws.cell(row=r, column=3, value=(
            '=IF($B{0}="","",IFERROR(VLOOKUP($B{0},目標タイム!$A:$B,2,FALSE),""))'
        ).format(r))
        ws.cell(row=r, column=5, value=(
            '=IF(OR($C{0}="",$D{0}=""),"",$D{0}-$C{0})').format(r))
        ws.cell(row=r, column=6, value=(
            '=IF($B{0}="","",COUNTIF($B$5:$B{0},$B{0}))').format(r))
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.border = BOX
            if col in (3, 4, 5, 6):
                cell.alignment = Alignment(horizontal='center')
            if r % 2 == 0:
                cell.fill = GREY
        dv.add(ws.cell(row=r, column=2))

    rng = 'E5:E%d' % (4 + ROWS)
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='lessThanOrEqual', formula=['0'], fill=GREEN))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='greaterThan', formula=['0'], fill=RED))
    ws.freeze_panes = 'A5'

    # ---------------- 集計 ----------------
    ws3 = wb.create_sheet('集計')
    ws3['A1'] = '種目ごとの記録'
    ws3['A1'].font = Font(bold=True, size=15, color=INK)
    head(ws3, 3, ['種目', '目標', '回数', '最速', '平均', '直近', '達成'],
         [22, 9, 8, 9, 9, 9, 10])
    for i, (nm, mn, _) in enumerate(MENU):
        r = 4 + i
        ws3.cell(row=r, column=1, value=nm)
        ws3.cell(row=r, column=2, value=mn)
        ws3.cell(row=r, column=3,
                 value='=COUNTIFS(記録!$B:$B,$A%d,記録!$D:$D,">0")' % r)
        ws3.cell(row=r, column=4,
                 value='=IF($C%d=0,"",MINIFS(記録!$D:$D,記録!$B:$B,$A%d))'
                       % (r, r))
        ws3.cell(row=r, column=5,
                 value='=IF($C{0}=0,"",ROUND(AVERAGEIFS(記録!$D:$D,'
                       '記録!$B:$B,$A{0},記録!$D:$D,">0"),1))'.format(r))
        ws3.cell(row=r, column=6,
                 value='=IF($C{0}=0,"",LOOKUP(2,1/((記録!$B$5:$B$204=$A{0})'
                       '*(記録!$D$5:$D$204>0)),記録!$D$5:$D$204))'.format(r))
        ws3.cell(row=r, column=7,
                 value='=IF($C{0}=0,"—",IF($F{0}<=$B{0},"達成","あと"&'
                       'ROUND($F{0}-$B{0},1)&"分"))'.format(r))
        for col in range(1, 8):
            c = ws3.cell(row=r, column=col)
            c.border = BOX
            if col > 1:
                c.alignment = Alignment(horizontal='center')
    n = len(MENU)
    ws3.conditional_formatting.add('G4:G%d' % (3 + n), CellIsRule(
        operator='equal', formula=['"達成"'], fill=GREEN))

    ch = BarChart()
    ch.type = 'col'
    ch.title = '目標 と 直近タイム'
    ch.height, ch.width = 9, 20
    data = Reference(ws3, min_col=2, max_col=2, min_row=3, max_row=3 + n - 1)
    data2 = Reference(ws3, min_col=6, max_col=6, min_row=3, max_row=3 + n - 1)
    cats = Reference(ws3, min_col=1, min_row=4, max_row=3 + n - 1)
    ch.add_data(data, titles_from_data=True)
    ch.add_data(data2, titles_from_data=True)
    ch.set_categories(cats)
    ws3.add_chart(ch, 'I3')

    # ---------------- 本番の流れ ----------------
    ws4 = wb.create_sheet('本番の流れ')
    ws4['A1'] = '本番 5時間の流れ（2026年9月13日 11:00〜16:00）'
    ws4['A1'].font = Font(bold=True, size=15, color=INK)
    head(ws4, 3, ['時刻', 'やること', '分', '終わっていないと危ないライン'],
         [10, 26, 8, 44])
    flow = [
        ('11:00', '読み取り・条件の書き出し', 25, '11:25までに条件表が埋まっている'),
        ('11:25', 'エスキス', 60, '12:25までに間取りが決まっている'),
        ('12:25', '記述（先に片づける）', 25, '12:50までに記述は全部埋める'),
        ('12:50', '1階平面図 兼 配置図', 40, '13:30'),
        ('13:30', '2階平面図', 30, '14:00'),
        ('14:00', '3階平面図', 25, '14:25'),
        ('14:25', '床伏図 兼 小屋伏図', 35, '15:00'),
        ('15:00', '立面図', 20, '15:20'),
        ('15:20', '部分詳細図 1/20', 25, '15:45'),
        ('15:45', '最終確認（ランクⅣの6項目）', 15, '16:00 終了'),
    ]
    for i, row in enumerate(flow):
        r = 4 + i
        for col, v in enumerate(row, start=1):
            c = ws4.cell(row=r, column=col, value=v)
            c.border = BOX
            if col in (1, 3):
                c.alignment = Alignment(horizontal='center')
        if i == 2:
            for col in range(1, 5):
                ws4.cell(row=r, column=col).fill = PatternFill(
                    'solid', fgColor='FFF3C4')
    ws4.cell(row=16, column=1,
             value='※ 記述を先にやるのは、あとに回すと必ず時間が足りなくなるから。'
                   '図面は雑でも点が残るが、記述の白紙はまるごと失点。')
    ws4.cell(row=16, column=1).font = Font(size=10, color='6B6459')
    ws4.cell(row=18, column=1, value='最後の15分でみるところ（ランクⅣ回避）')
    ws4.cell(row=18, column=1).font = Font(bold=True, size=12, color=ACC)
    checks = [
        '木造3階建てになっているか（構造・階数の指定どおりか）',
        '図面6枚＋面積表＋記述、⑴〜⑻すべて描き終わっているか（1つでも未完成なら不合格）',
        '平面図と伏図で柱の位置が合っているか',
        '面積表の数字が条件に合っているか',
        '要求された部屋がすべてあるか・階を間違えていないか',
        '各階に階段があるか／全部の部屋に入れるか',
        '室名・寸法・方位・縮尺を書き忘れていないか',
    ]
    for i, t in enumerate(checks):
        c = ws4.cell(row=19 + i, column=1, value='□　' + t)
        c.font = Font(size=11)
    return wb


if __name__ == '__main__':
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..',
                       'timetrial.xlsx')
    build().save(out)
    print('wrote', os.path.normpath(out))
